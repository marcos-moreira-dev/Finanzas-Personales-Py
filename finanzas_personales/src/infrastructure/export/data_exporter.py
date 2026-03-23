"""
Sistema de Exportación Avanzada - Infraestructura

Proporciona exportación a múltiples formatos:
- CSV (básico y avanzado)
- Excel (.xlsx) con múltiples hojas y gráficos
- JSON (para integración con otros sistemas)
- Backup completo de la base de datos

Características profesionales:
- Formato condicional en Excel
- Fórmulas automáticas
- Gráficos integrados
- Múltiples hojas organizadas
- Estilos corporativos

Ejemplo:
    >>> exporter = DataExporter()
    >>> exporter.exportar_excel_completo(
    ...     persona_id=1,
    ...     output_path="mi_finanzas.xlsx"
    ... )
"""
import csv
import json
import sqlite3
import logging
import shutil
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from pathlib import Path

import pandas as pd

# Intentar importar openpyxl para Excel
# Si no está disponible, se usa solo CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class DataExporter:
    """
    Exportador de datos a múltiples formatos.
    
    Permite exportar movimientos, resúmenes y reportes completos
    en formatos estándar para análisis externo o respaldo.
    
    Attributes:
        export_dir: Directorio donde se guardarán los archivos
    """
    
    def __init__(self, export_dir: Optional[str] = None):
        """
        Inicializa el exportador.
        
        Args:
            export_dir: Directorio de exportación (opcional)
        """
        if export_dir is None:
            # Usar directorio por defecto
            from ...infrastructure.config.settings import Config
            self.export_dir = Path(Config.EXPORT_DIR)
        else:
            self.export_dir = Path(export_dir)
        
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"DataExporter inicializado. Directorio: {self.export_dir}")
    
    def exportar_movimientos_csv(
        self,
        movimientos: List[Dict],
        output_path: Optional[str] = None,
        incluir_headers: bool = True
    ) -> str:
        """
        Exporta movimientos a CSV.
        
        Args:
            movimientos: Lista de diccionarios con movimientos
            output_path: Ruta de salida (opcional)
            incluir_headers: Si True, incluye fila de encabezados
            
        Returns:
            Ruta al archivo CSV generado
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.export_dir / f"movimientos_{timestamp}.csv"
        else:
            output_path = Path(output_path)
        
        logger.info(f"Exportando {len(movimientos)} movimientos a CSV: {output_path}")
        
        try:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                if movimientos:
                    # Obtener campos del primer registro
                    fieldnames = list(movimientos[0].keys())
                    
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    
                    if incluir_headers:
                        writer.writeheader()
                    
                    writer.writerows(movimientos)
                else:
                    f.write("No hay datos para exportar\n")
            
            logger.info(f"CSV exportado exitosamente: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error exportando CSV: {e}")
            raise
    
    def exportar_excel_completo(
        self,
        persona_id: int,
        persona_nombre: str,
        movimientos: List[Dict],
        resumen: Dict,
        output_path: Optional[str] = None
    ) -> str:
        """
        Exporta un reporte completo a Excel con múltiples hojas.
        
        Hojas generadas:
        1. Resumen - Totales y estadísticas
        2. Movimientos - Detalle completo
        3. Análisis - Gráficos y análisis
        
        Args:
            persona_id: ID de la persona
            persona_nombre: Nombre completo
            movimientos: Lista de movimientos
            resumen: Datos de resumen
            output_path: Ruta de salida (opcional)
            
        Returns:
            Ruta al archivo Excel
        """
        if not EXCEL_AVAILABLE:
            logger.warning("openpyxl no está instalado. Usando CSV como alternativa.")
            return self.exportar_movimientos_csv(movimientos, output_path)
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = persona_nombre.replace(" ", "_")[:20]
            output_path = self.export_dir / f"finanzas_{safe_name}_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        logger.info(f"Exportando reporte Excel: {output_path}")
        
        # Crear workbook
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hojas
        self._crear_hoja_resumen(wb, persona_nombre, resumen)
        self._crear_hoja_movimientos(wb, movimientos)
        self._crear_hoja_analisis(wb, resumen)
        
        # Guardar
        wb.save(output_path)
        logger.info(f"Excel exportado exitosamente: {output_path}")
        
        return str(output_path)
    
    def _crear_hoja_resumen(self, wb: Workbook, persona_nombre: str, resumen: Dict):
        """Crea la hoja de resumen."""
        ws = wb.create_sheet("Resumen", 0)
        
        # Estilos
        title_font = Font(bold=True, size=16, color="2C3E50")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        positive_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        negative_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Título
        ws['A1'] = f"Reporte Financiero - {persona_nombre}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, size=9)
        
        # Espacio
        ws.append([])
        
        # Totales
        ws['A4'] = "CONCEPTO"
        ws['B4'] = "VALOR"
        ws['A4'].font = header_font
        ws['B4'].font = header_font
        ws['A4'].fill = header_fill
        ws['B4'].fill = header_fill
        
        row = 5
        ws[f'A{row}'] = "Total Ingresos"
        ws[f'B{row}'] = resumen.get('total_ingresos', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color="27AE60")
        row += 1
        
        ws[f'A{row}'] = "Total Gastos"
        ws[f'B{row}'] = resumen.get('total_gastos', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, color="E74C3C")
        row += 1
        
        ws[f'A{row}'] = "Saldo Neto"
        ws[f'B{row}'] = resumen.get('saldo', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _crear_hoja_movimientos(self, wb: Workbook, movimientos: List[Dict]):
        """Crea la hoja de movimientos."""
        ws = wb.create_sheet("Movimientos")
        
        # Headers
        headers = ['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']
        ws.append(headers)
        
        # Estilo headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for mov in movimientos:
            ws.append([
                mov.get('fecha', ''),
                mov.get('tipo', ''),
                mov.get('categoria', ''),
                mov.get('descripcion', ''),
                mov.get('monto', 0)
            ])
        
        # Formato de moneda en columna E
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '$#,##0.00'
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        cell.font = Font(color="E74C3C")
                    else:
                        cell.font = Font(color="27AE60")
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
    
    def _crear_hoja_analisis(self, wb: Workbook, resumen: Dict):
        """Crea la hoja de análisis con gráficos."""
        ws = wb.create_sheet("Análisis")
        
        ws['A1'] = "Análisis de Gastos por Categoría"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Datos para gráfico
        if 'gastos_por_categoria' in resumen:
            ws['A3'] = "Categoría"
            ws['B3'] = "Monto"
            
            row = 4
            for cat in resumen['gastos_por_categoria'][:10]:  # Top 10
                ws[f'A{row}'] = cat['nombre']
                ws[f'B{row}'] = cat['total']
                row += 1
            
            # Crear gráfico de pastel
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=4, max_row=row-1)
            data = Reference(ws, min_col=2, min_row=3, max_row=row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución de Gastos"
            
            ws.add_chart(pie, "D3")
    
    def exportar_json(
        self,
        datos: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> str:
        """
        Exporta datos a formato JSON.
        
        Útil para integración con otros sistemas o backups estructurados.
        
        Args:
            datos: Diccionario con datos a exportar
            output_path: Ruta de salida (opcional)
            
        Returns:
            Ruta al archivo JSON
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.export_dir / f"export_{timestamp}.json"
        else:
            output_path = Path(output_path)
        
        logger.info(f"Exportando a JSON: {output_path}")
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(datos, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON exportado: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error exportando JSON: {e}")
            raise
    
    def crear_backup_completo(
        self,
        db_path: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        Crea un backup completo de la base de datos.
        
        Args:
            db_path: Ruta a la base de datos SQLite
            output_path: Ruta de salida (opcional)
            
        Returns:
            Ruta al archivo de backup
        """
        db_path = Path(db_path)
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.export_dir / f"backup_{timestamp}.db"
        else:
            output_path = Path(output_path)
        
        logger.info(f"Creando backup de BD: {db_path} -> {output_path}")
        
        try:
            shutil.copy2(db_path, output_path)
            logger.info(f"Backup creado exitosamente: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error creando backup: {e}")
            raise
    
    def listar_exports(self) -> List[Dict]:
        """
        Lista todos los archivos exportados.
        
        Returns:
            Lista de diccionarios con información de cada exportación
        """
        exports = []
        
        for file_path in self.export_dir.iterdir():
            if file_path.is_file():
                stat = file_path.stat()
                exports.append({
                    'nombre': file_path.name,
                    'ruta': str(file_path),
                    'tipo': file_path.suffix,
                    'tamaño': stat.st_size,
                    'fecha': datetime.fromtimestamp(stat.st_mtime)
                })
        
        # Ordenar por fecha descendente
        exports.sort(key=lambda x: x['fecha'], reverse=True)
        
        return exports
